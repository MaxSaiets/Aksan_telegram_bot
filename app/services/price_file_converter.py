"""
Filter an xlsx file: keep only rows where any cell has a non-default
background fill OR font color (i.e. rows the user manually highlighted).
"""
from __future__ import annotations

from pathlib import Path

from openpyxl import load_workbook, Workbook
from openpyxl.styles.fills import GradientFill, PatternFill
from openpyxl.utils import get_column_letter

from app.utils.logger import get_logger

logger = get_logger(__name__)

# RGB values that mean "no color" / default
_NO_COLOR_RGB = {"00000000", "FFFFFFFF", "FF000000"}
# Indexed color 64 = "automatic" (no explicit color set)
_AUTO_INDEXED = {64}
# Theme indices that are background/text defaults (white bg, black text)
_DEFAULT_THEME = {0, 1}


def _color_is_set(color) -> bool:
    """Return True if this Color object represents a real non-default color."""
    if color is None:
        return False
    t = getattr(color, "type", None)
    if t == "theme":
        theme_idx = getattr(color, "theme", None)
        return theme_idx not in _DEFAULT_THEME
    if t == "indexed":
        idx = getattr(color, "indexed", 64)
        return idx not in _AUTO_INDEXED
    rgb = (getattr(color, "rgb", None) or "").upper()
    return bool(rgb) and rgb not in _NO_COLOR_RGB


def _fill_is_colored(fill) -> bool:
    if fill is None:
        return False
    if isinstance(fill, GradientFill):
        return True
    if not isinstance(fill, PatternFill):
        return False
    if fill.patternType in (None, "none"):
        return False
    # patternType is "solid" or any other pattern — check foreground color
    return _color_is_set(fill.fgColor) or _color_is_set(fill.bgColor)


def _font_is_colored(font) -> bool:
    if font is None:
        return False
    return _color_is_set(getattr(font, "color", None))


def _row_is_highlighted(ws, row_idx: int) -> bool:
    for cell in ws[row_idx]:
        if _fill_is_colored(cell.fill):
            return True
        if _font_is_colored(cell.font):
            return True
    return False


def _debug_row(ws, row_idx: int) -> str:
    parts = []
    for cell in ws[row_idx]:
        fill = cell.fill
        font = cell.font
        f_info = ""
        if isinstance(fill, PatternFill):
            fg = fill.fgColor
            f_info = f"fill=pat/{fill.patternType}/fg.type={getattr(fg,'type','?')}/rgb={getattr(fg,'rgb','?')}/theme={getattr(fg,'theme','?')}/idx={getattr(fg,'indexed','?')}"
        elif isinstance(fill, GradientFill):
            f_info = "fill=gradient"
        fc = getattr(font, "color", None) if font else None
        fnt_info = f"font.color.type={getattr(fc,'type','?')}/rgb={getattr(fc,'rgb','?')}" if fc else "font.color=None"
        parts.append(f"[{cell.coordinate}: {f_info} | {fnt_info}]")
    return " ".join(parts)


def filter_colored_rows(input_path: Path, output_path: Path) -> tuple[int, str]:
    """
    Read input_path xlsx, write only highlighted rows to output_path.
    Returns (kept_count, debug_info).
    """
    wb_in = load_workbook(input_path)
    ws_in = wb_in.active

    wb_out = Workbook()
    ws_out = wb_out.active
    ws_out.title = ws_in.title

    max_col = ws_in.max_column
    header_row = [ws_in.cell(1, c).value for c in range(1, max_col + 1)]
    ws_out.append(header_row)

    kept = 0
    debug_sample = ""
    for row_idx in range(2, ws_in.max_row + 1):
        highlighted = _row_is_highlighted(ws_in, row_idx)
        # Capture debug info for first 3 data rows
        if row_idx <= 4:
            flag = "✓" if highlighted else "✗"
            debug_sample += f"\nRow {row_idx} {flag}: {_debug_row(ws_in, row_idx)}"
        if highlighted:
            values = [ws_in.cell(row_idx, c).value for c in range(1, max_col + 1)]
            ws_out.append(values)
            kept += 1

    for col_idx in range(1, max_col + 1):
        letter = get_column_letter(col_idx)
        max_len = len(str(header_row[col_idx - 1] or ""))
        for row_idx in range(2, ws_out.max_row + 1):
            val = ws_out.cell(row_idx, col_idx).value
            max_len = max(max_len, len(str(val or "")))
        ws_out.column_dimensions[letter].width = min(max_len + 4, 60)

    wb_out.save(str(output_path))
    logger.info("Converted: %d rows kept | debug: %s", kept, debug_sample)
    return kept, debug_sample
